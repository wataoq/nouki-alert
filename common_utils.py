# common_utils.py
# ---------------------------------------------------------------------------
# å…±æœ‰ãƒ¦ãƒ¼ãƒ†ã‚£ãƒªãƒ†ã‚£ï¼š
#   â€¢ Dropbox ã‹ã‚‰ Excel ã‚’å–å¾—ï¼ˆdownload_excelï¼‰
#   â€¢ è¡Œã‚¹ã‚­ãƒƒãƒ—åˆ¤å®šï¼šã‚»ãƒ«ã®èƒŒæ™¯è‰² or æ–‡å­—è‰²ãŒç™½ä»¥å¤–ãªã‚‰é™¤å¤–ï¼ˆrows_to_skip_by_colorï¼‰
#   â€¢ ã‚¢ãƒ©ãƒ¼ãƒˆåˆ¤å®šï¼ˆshould_alertï¼‰
#   â€¢ SMTP çµŒç”±ã§ãƒ¡ãƒ¼ãƒ«é€ä¿¡ï¼ˆsend_emailï¼‰
# ---------------------------------------------------------------------------
import os
import io
import datetime
import logging
import smtplib
from email.mime.text import MIMEText
from typing import Set

import dropbox
from openpyxl import load_workbook

IS_DRY_RUN = os.getenv("DRY_RUN", "0") == "1"   # â˜…è¿½åŠ 

logging.basicConfig(level=logging.INFO)

# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
# Dropbox
# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
def get_dropbox_client() -> dropbox.Dropbox:
    """ç’°å¢ƒå¤‰æ•°ã‹ã‚‰ãƒªãƒ•ãƒ¬ãƒƒã‚·ãƒ¥ãƒˆãƒ¼ã‚¯ãƒ³ã‚’èª­ã¿è¾¼ã‚“ã§ Dropbox ã‚¯ãƒ©ã‚¤ã‚¢ãƒ³ãƒˆã‚’è¿”ã™"""
    return dropbox.Dropbox(
        app_key              = os.environ["DROPBOX_APP_KEY"],
        app_secret           = os.environ["DROPBOX_APP_SECRET"],
        oauth2_refresh_token = os.environ["DROPBOX_REFRESH_TOKEN"],
        timeout              = int(os.getenv("DROPBOX_TIMEOUT", 900)),  # â˜…è¿½åŠ 
    )


def download_excel(path: str) -> bytes | None:
    """
    Dropbox ã‹ã‚‰æŒ‡å®šãƒ‘ã‚¹ã®ãƒ•ã‚¡ã‚¤ãƒ«ã‚’ãƒ€ã‚¦ãƒ³ãƒ­ãƒ¼ãƒ‰ã—ã¦ raw bytes ã‚’è¿”ã™ã€‚
    å¤±æ•—ã—ãŸã‚‰ None ã‚’è¿”ã™ã€‚
    """
    try:
        _, res = get_dropbox_client().files_download(path)
        logging.info("âœ…  Dropbox ã‹ã‚‰ Excel ã‚’å–å¾—: %s", path)
        return res.content
    except Exception as e:
        logging.error("âŒ Dropbox ãƒ€ã‚¦ãƒ³ãƒ­ãƒ¼ãƒ‰å¤±æ•—: %s", e)
        return None


# ------------------------------------------------------------------
#  â– ã“ã‚Œã ã‘ã§ OK
#     - èƒŒæ™¯è‰²ãŒ #F7DFDFï¼ˆARGB ã§ã‚‚ RGB ã§ã‚‚å¯ï¼‰ã®ã‚»ãƒ«ã‚’æŒã¤è¡Œã ã‘é™¤å¤–
#     - æ–‡å­—è‰²ã¯åˆ¤å®šã—ãªã„
# ------------------------------------------------------------------
SKIP_BG_HEX = {"f7dfdf"}        # é™¤å¤–ã—ãŸã„ 6 æ¡ RGB ã‚’åˆ—æŒ™

def _is_skip_color(argb: str | None) -> bool:
    """openpyxl ã® ARGB 8æ¡ or RGB 6æ¡ã‚’å—ã‘å–ã‚Šã€å¯¾è±¡è‰²ãªã‚‰ True"""
    if argb is None:
        return False
    return argb[-6:].upper() in SKIP_BG_HEX      # ä¸‹ 6 æ¡ã§æ¯”è¼ƒ

def rows_to_skip_by_color(raw_bytes: bytes, sheet_name: str,
                          target_col: int,
                          first_data_row_excel: int = 8) -> set[int]:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(raw_bytes), data_only=True)
    ws = wb[sheet_name]

    skip = set()
    for df_row_idx, row in enumerate(ws.iter_rows(min_row=first_data_row_excel), 0):
        cell = row[target_col]
        bg_rgb = getattr(cell.fill.fgColor, "rgb", None)
        if _is_skip_color(bg_rgb):
            skip.add(df_row_idx)
    return skip


# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
# SMTP ãƒ¡ãƒ¼ãƒ«é€ä¿¡
# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
def send_email(subject: str, body: str):
    """
    TEXT ãƒ¡ãƒ¼ãƒ«ã‚’ SMTP ã§é€ä¿¡ã€‚
    DRY_RUN=1 ã®å ´åˆã¯ãƒ­ã‚°å‡ºåŠ›ã ã‘ã§ã‚¹ã‚­ãƒƒãƒ—ã€‚
    """

    if IS_DRY_RUN:                                 # â˜…è¿½åŠ 
        logging.info("ğŸŸ¡ DRYâ€‘RUN â†’ ãƒ¡ãƒ¼ãƒ«é€ä¿¡ã‚¹ã‚­ãƒƒãƒ—: %s", subject)
        return

    smtp_server   = os.environ["SMTP_SERVER"]
    smtp_port     = int(os.environ.get("SMTP_PORT", 587))
    smtp_user     = os.environ["SMTP_USER"]
    smtp_password = os.environ["SMTP_PASSWORD"]
    recipients    = os.environ["EMAIL_RECIPIENTS"].split(",")

    msg = MIMEText(body, "plain", "utf-8")
    msg["Subject"] = subject
    msg["From"]    = smtp_user
    msg["To"]      = ", ".join(recipients)

    try:
        with smtplib.SMTP(smtp_server, smtp_port) as s:
            s.starttls()
            s.login(smtp_user, smtp_password)
            s.send_message(msg)
        logging.info("âœ… ãƒ¡ãƒ¼ãƒ«é€ä¿¡å®Œäº† â†’ %s", recipients)
    except Exception as e:
        logging.error("âŒ ãƒ¡ãƒ¼ãƒ«é€ä¿¡ã‚¨ãƒ©ãƒ¼: %s", e)
        raise



# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
# ã‚¢ãƒ©ãƒ¼ãƒˆåˆ¤å®šï¼ˆå…±é€šãƒ­ã‚¸ãƒƒã‚¯ï¼‰
# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
def should_alert(due: datetime.date, alert_days: int) -> bool:
    """
    é€šçŸ¥åˆ¤å®š:
      â€¢ æŒ‡å®šæ—¥å‰ (alert_days) ã®ã¿é€šçŸ¥
      â€¢ 1ï½2æ—¥é…å»¶ã®ã‚‚ã®ã¯ âš ï¸ é€šçŸ¥
      â€¢ 3æ—¥ä»¥ä¸Šé…å»¶ã—ãŸã‚‚ã®ã¯ç„¡è¦–
    """
    today = datetime.date.today()
    delta = (due - today).days  # æœªæ¥ = æ­£ã€éå» = è² 

    if delta == alert_days:
        return True   # æŒ‡å®šæ—¥å‰ã´ã£ãŸã‚Š
    if -2 <= delta < 0:
        return True   # é…å»¶ 1ï½2 æ—¥
    return False
